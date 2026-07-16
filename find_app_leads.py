"""
Szuka potencjalnych klientow dla agencji UI/reklamowej: aplikacji z App Store,
ktore maja mala liczbe ocen (proxy dla malej liczby uzytkownikow).

Tym razem SZEROKO:
- wiele storefrontow (krajow) App Store na raz, nie tylko Polska
- wiele ogolnych, "niszowych/gimmickowych" fraz (kalkulator z kotami,
  dzwieki, generatory itp.), nie tylko appki dla lokalnych firm
- cel: ~500 leadow w jednym pliku Excel

Zrodlo danych: publiczne iTunes Search API (bez klucza).
Nie daje dokladnej liczby instalacji ani nie ocenia estetyki automatycznie
- to pierwsze sito. Ostateczna ocena "czy jest brzydkie" i tak wymaga
rzutu oka na screeny w App Store (link jest w kolumnie ostatniej).

Uzycie:
    pip install requests openpyxl
    python find_app_leads.py

Uwaga o czasie dzialania: przy ~30 frazach x ~20 krajow to nawet
kilkaset zapytan do API. Zeby nie zostac zblokowanym, skrypt czeka
mniej wiecej 1,5 sek. miedzy zapytaniami - liczy sie wiec na kilkanascie
minut dzialania. To normalne, nie przerywaj.
"""

import random
import time
from datetime import datetime, timezone
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- KONFIGURACJA -------------------------------------------------------

# Ogolne, "gimmickowe"/niszowe frazy - typ appek robionych przez jedna
# osobe po godzinach, bez budzetu na design. Dopisz swoje pomysly.
TERMS = [
    "cat calculator",
    "kawaii calculator",
    "unicorn calculator",
    "cute alarm clock",
    "fart sounds app",
    "burp sounds app",
    "dog sound board",
    "cute weather widget",
    "meme sound board",
    "random name picker",
    "decision wheel spinner",
    "coin flip app",
    "dice roller app",
    "cute countdown timer",
    "cute habit tracker",
    "mood diary app",
    "quote of the day app",
    "daily horoscope app",
    "simple period tracker",
    "water reminder app",
    "white noise sleep sounds",
    "baby name generator",
    "pet name generator",
    "zodiac compatibility app",
    "tarot card reader app",
    "cute notes widget",
    "photo collage cute",
    "sticker maker app",
    "simple flashlight app",
    "cute tip calculator",
    "simple workout timer",
    "magic 8 ball app",
    "cute journal app",
    "rock paper scissors app",
]

# Storefronty App Store do przeszukania (kody krajow ISO).
COUNTRIES = [
    "us", "gb", "ca", "au", "nz", "ie",
    "de", "fr", "es", "it", "nl", "se", "no", "dk", "fi",
    "jp", "kr", "br", "mx", "pl",
]

MAX_RATING_COUNT = 200      # prog "malo uzytkownikow" (szeroki, przycinamy pozniej)
RESULTS_PER_CALL = 200      # max dozwolony przez iTunes Search API
TARGET_LEADS = 500          # ile leadow finalnie chcemy w Excelu
EARLY_STOP_AT = int(TARGET_LEADS * 1.5)  # zbieramy troche zapasu, potem przycinamy
SLEEP_BETWEEN_CALLS = 1.5   # sekundy, zeby nie zostac zrzucanym przez limit API

OUTPUT_XLSX = "app_leads.xlsx"

# --- LOGIKA ---------------------------------------------------------------

def search_apps(term: str, country: str, limit: int = RESULTS_PER_CALL):
    url = "https://itunes.apple.com/search"
    params = {"term": term, "country": country, "entity": "software", "limit": limit}
    r = requests.get(url, params=params, timeout=20)
    if r.status_code != 200:
        raise requests.RequestException(f"HTTP {r.status_code}")
    return r.json().get("results", [])


def days_since_update(app: dict) -> Optional[int]:
    release_date_str = app.get("currentVersionReleaseDate")
    if not release_date_str:
        return None
    release_date = datetime.fromisoformat(release_date_str.replace("Z", "+00:00"))
    return (datetime.now(timezone.utc) - release_date).days


def is_candidate(app: dict) -> bool:
    rating_count = app.get("userRatingCount", 0) or 0
    return rating_count <= MAX_RATING_COUNT


def write_excel(candidates, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leady"

    headers = [
        "Nazwa", "Deweloper", "Liczba ocen", "Kategoria",
        "Kraj (storefront)", "Ostatnia aktualizacja", "Link do App Store",
    ]

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    band_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    for row_idx, c in enumerate(candidates, start=2):
        values = [
            c["nazwa"], c["deweloper"], c["liczba_ocen"], c["kategoria"],
            c["kraj"], c["ostatnia_aktualizacja"], c["url"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="center" if col_idx in (3, 5, 6) else "left")
            if row_idx % 2 == 0:
                cell.fill = band_fill

        url_cell = ws.cell(row=row_idx, column=7)
        url_cell.hyperlink = c["url"]
        url_cell.font = Font(name="Arial", size=10, color="1155CC", underline="single")

    widths = [32, 24, 12, 20, 14, 18, 48]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(candidates) + 1}"
    wb.save(path)


def main():
    combos = [(term, country) for term in TERMS for country in COUNTRIES]
    random.shuffle(combos)  # zeby nie wyczerpac limitu na jednej frazie/kraju

    seen_ids = set()
    candidates = []

    for i, (term, country) in enumerate(combos, start=1):
        print(f"[{i}/{len(combos)}] {term} ({country}) - znaleziono dotad: {len(candidates)}")
        try:
            results = search_apps(term, country)
        except requests.RequestException as e:
            print(f"  blad zapytania, czekam dluzej i pomijam: {e}")
            time.sleep(5)
            continue

        for app in results:
            app_id = app.get("trackId")
            if app_id in seen_ids:
                continue
            seen_ids.add(app_id)

            if is_candidate(app):
                candidates.append({
                    "nazwa": app.get("trackName"),
                    "deweloper": app.get("sellerName"),
                    "liczba_ocen": app.get("userRatingCount", 0),
                    "kategoria": app.get("primaryGenreName"),
                    "kraj": country.upper(),
                    "ostatnia_aktualizacja": app.get("currentVersionReleaseDate", "")[:10],
                    "url": app.get("trackViewUrl"),
                })

        time.sleep(SLEEP_BETWEEN_CALLS)

        if len(candidates) >= EARLY_STOP_AT:
            print(f"\nOsiagnieto zapas {EARLY_STOP_AT} leadow, konczymy zbieranie wczesniej.")
            break

    candidates.sort(key=lambda c: c["liczba_ocen"])
    final_candidates = candidates[:TARGET_LEADS]

    if not final_candidates:
        print("\nNie znaleziono zadnych pasujacych aplikacji. Sprobuj zmienic frazy w TERMS.")
        return

    write_excel(final_candidates, OUTPUT_XLSX)
    print(f"\nZapisano {len(final_candidates)} leadow do {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
